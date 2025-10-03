# gui/windows/edit_window.py
import tkinter as tk
from tkinter import ttk, messagebox
from typing import Dict, Any

# Импорты из проекта
from core.database import find_client
from config.settings import WINDOW_WIDTH, ENTRY_WIDTH


def open_edit_window(parent):
    """
    Окно для редактирования данных клиента по VIN
    :param parent: родительское окно
    """
    window = tk.Toplevel(parent)
    window.title("✏️ Редактирование данных")
    window.geometry(f"{WINDOW_WIDTH}x600")
    window.resizable(False, False)
    window.transient(parent)
    window.grab_set()

    # Поля формы (соответствуют базе)
    fields_config = [
        ("Фамилия", "Фамилия"),
        ("Имя", "Имя"),
        ("Отчество", "Отчество"),
        ("Марка авто", "Марка авто"),
        ("VIN", "VIN"),
        ("Индекс", "Индекс"),
        ("Адрес", "Адрес"),
        ("Паспорт (серия и номер)", "Паспорт (серия и номер)"),
        ("Кем выдан", "Кем выдан"),
        ("Дата выдачи", "Дата выдачи"),
        ("Код подразделения", "Код подразделения"),
        ("Телефон", "Телефон"),
        ("Дата рождения", "Дата рождения"),
    ]

    # --- 1. Поле поиска ---
    search_frame = ttk.Frame(window)
    search_frame.pack(pady=15, padx=20, fill="x")

    ttk.Label(search_frame, text="Введите VIN для редактирования:").pack(side="left")
    search_entry = ttk.Entry(search_frame, width=ENTRY_WIDTH)
    search_entry.pack(side="left", padx=10)
    search_entry.focus()

    entries = {}
    client_row_num = None
    original_data = {}

    def load_client():
        nonlocal client_row_num, original_data
        vin = search_entry.get().strip()
        if not vin:
            messagebox.showwarning("Внимание", "Введите VIN.")
            return

        # Поиск по VIN
        client_data = find_client(vin)
        if client_data is None:
            messagebox.showerror("Ошибка", "Клиент с таким VIN не найден.")
            return

        client_row_num = int(client_data["№"]) + 1  # Excel-строка (учитывая заголовок)
        original_data = client_data.to_dict()

        # Заполнение полей
        for label_text, field_name in fields_config:
            entry = entries[field_name]  # Используем имя поля как ключ
            value = original_data.get(field_name, "")
            entry.delete(0, tk.END)
            entry.insert(0, str(value))

        # Разблокировать кнопку сохранения
        save_btn.config(state="normal")

    def save_changes():
        if client_row_num is None:
            return

        # Сбор новых данных
        new_data = {}
        for label_text, field_name in fields_config:
            new_data[field_name] = entries[field_name].get().strip()

        # Генерация имени папки
        folder_name = f"{new_data['Фамилия']}_{new_data['Марка авто']}_vin {new_data['VIN']}_{new_data['Индекс']}"

        # Обновление в Excel
        try:
            from openpyxl import load_workbook
            from config.paths import CLIENTS_DB_PATH

            wb = load_workbook(CLIENTS_DB_PATH)
            ws = wb["Folder"]

            # Все поля в правильном порядке
            ordered_fields = [
                "№", "Фамилия", "Имя", "Отчество", "Марка авто", "VIN",
                "Индекс", "Папка", "Адрес", "Паспорт (серия и номер)",
                "Кем выдан", "Дата выдачи", "Код подразделения",
                "Телефон", "Дата рождения", "Дата создания папки"
            ]

            row_data = [
                original_data["№"],  # № не меняем
                new_data["Фамилия"],
                new_data["Имя"],
                new_data["Отчество"],
                new_data["Марка авто"],
                new_data["VIN"],
                new_data["Индекс"],
                folder_name,
                new_data["Адрес"],
                new_data["Паспорт (серия и номер)"],
                new_data["Кем выдан"],
                new_data["Дата выдачи"],
                new_data["Код подразделения"],
                new_data["Телефон"],
                new_data["Дата рождения"],
                original_data["Дата создания папки"]  # дата создания не меняется
            ]

            # Перезапись строки
            for col, value in enumerate(row_data, 1):
                ws.cell(row=client_row_num, column=col, value=value)

            wb.save(CLIENTS_DB_PATH)
            messagebox.showinfo("Успех", "Данные успешно обновлены!")
            window.destroy()

        except Exception as e:
            messagebox.showerror("Ошибка", f"Не удалось сохранить изменения:\n{e}")

    # --- 2. Форма редактирования ---
    form_frame = ttk.Frame(window)
    form_frame.pack(pady=10, padx=20, fill="both", expand=True)

    for i, (label_text, field_name) in enumerate(fields_config):
        ttk.Label(form_frame, text=label_text + ":").grid(
            row=i, column=0, sticky="e", padx=(5, 5), pady=5
        )
        entry = ttk.Entry(form_frame, width=ENTRY_WIDTH)
        entry.grid(row=i, column=1, padx=(0, 10), pady=5, sticky="ew")
        entries[field_name] = entry

    # Кнопка поиска
    ttk.Button(window, text="🔍 Найти клиента", command=load_client).pack(pady=5)

    # Кнопки управления
    button_frame = ttk.Frame(window)
    button_frame.pack(pady=15)

    ttk.Button(button_frame, text="Отмена", command=window.destroy).pack(side="left", padx=5)
    save_btn = ttk.Button(button_frame, text="✅ Сохранить изменения", command=save_changes, state="disabled")
    save_btn.pack(side="left", padx=5)

    # Обработка Enter
    search_entry.bind("<Return>", lambda event: load_client())
