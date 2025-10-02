# core/database.py
import logging
import pandas as pd
from openpyxl import load_workbook
from typing import Optional, Dict, Any

# Импортируем пути
from config.paths import CLIENTS_DB_PATH, CONTRACTS_DB_PATH


def get_next_client_id(sheet_name="Folder") -> int:
    """Возвращает следующий номер клиента (№)"""
    try:
        df = pd.read_excel(CLIENTS_DB_PATH, sheet_name=sheet_name)
        return int(df["№"].max()) + 1 if not df.empty else 1
    except Exception as e:
        print(f"Ошибка чтения ID: {e}")
        return 1


def find_client(search_term: str) -> Optional[pd.Series]:
    """
    Ищет клиента по VIN, ФИО
    Возвращает строку DataFrame или None
    """
    try:
        df = pd.read_excel(CLIENTS_DB_PATH, sheet_name="Folder")
        df = df.fillna("")  # Заменяем NaN

        mask = (
            df["VIN"].astype(str).str.contains(search_term, case=False, na=False) |
            df["Фамилия"].str.contains(search_term, case=False, na=False) |
            df["Имя"].str.contains(search_term, case=False, na=False) |
            df["Отчество"].str.contains(search_term, case=False, na=False)
        )
        if mask.any():
            return df[mask].iloc[0]
    except Exception as e:
        print(f"Ошибка поиска клиента: {e}")
    return None


def save_client(data: Dict[str, Any]) -> bool:
    """Сохраняет нового клиента в Excel"""
    try:
        wb = load_workbook(CLIENTS_DB_PATH)
        ws = wb["Folder"]
        ws.append(list(data.values()))
        wb.save(CLIENTS_DB_PATH)
        logging.info(f"Клиент сохранён: {data['Фамилия']} {data['Имя']}")
        return True
    except Exception as e:
        logging.error(f"Ошибка сохранения клиента: {e}")
        return False


def get_next_contract_number() -> str:
    """
    Генерирует следующий номер договора: 101-ИП, 102-ИП и т.д.
    Основывается на максимальном номере в столбце 'Номер договора'
    """
    try:
        df = pd.read_excel(CONTRACTS_DB_PATH, sheet_name="Registry")
        if df.empty:
            return "101-ИП"

        # Извлекаем числовую часть из 'Номер договора' (например, из '101-ИП' → 101)
        numbers = []
        for num_str in df["Номер договора"].dropna():  # Исключаем NaN
            if isinstance(num_str, str) and "-ИП" in num_str:
                try:
                    n = int(num_str.replace("-ИП", "").strip())
                    numbers.append(n)
                except ValueError:
                    continue  # Пропускаем некорректные значения

        last_num = max(numbers) if numbers else 100
        return f"{last_num + 1}-ИП"

    except Exception as e:
        logging.warning(f"⚠️ Ошибка при получении номера договора: {e}")
        return "101-ИП"


def save_contract_record(contract_data: Dict[str, Any]) -> bool:
    """Сохраняет запись о договоре в реестр"""
    try:
        wb = load_workbook(CONTRACTS_DB_PATH)
        ws = wb["Registry"]
        row = [
            contract_data["Номер"],
            contract_data["ФИО"],
            contract_data["Номер договора"],
            contract_data["Телефон"],
            contract_data["Индекс"],
            contract_data["Дата"]
        ]
        ws.append(row)
        wb.save(CONTRACTS_DB_PATH)
        logging.info(f"Договор сохранён: {contract_data['Номер договора']}")
        return True
    except Exception as e:
        logging.error(f"Ошибка сохранения договора: {e}")
        return False


def get_next_registry_id() -> int:
    """
    Возвращает следующий порядковый номер для реестра договоров
    """
    try:
        df = pd.read_excel(CONTRACTS_DB_PATH, sheet_name="Registry")
        if df.empty:
            return 1
        last_id = df["Номер"].max()
        return int(last_id) + 1
    except Exception as e:
        logging.warning(f"⚠️ Не удалось прочитать Номер из реестра: {e}")
        return 1


# core/database.py

def get_client_data_for_contract(search_term: str) -> list:
    """
    Возвращает список данных клиента в порядке, соответствующем расположению & в шаблоне.
    Используется для совместимости с шаблонами, где плейсхолдер — символ &
    """
    client = find_client(search_term)
    if client is None:
        logging.warning(f"Клиент не найден по запросу: {search_term}")
        return []

    # Форматируем телефон
    # phone, index_part = format_phone(client["Телефон"])

    # Формируем ФИО и инициалы
    full_fio = f"{client['Фамилия']} {client['Имя']} {client['Отчество']}"
    short_fio = f"{client['Фамилия']} {client['Имя'][0]}. {client['Отчество'][0]}."

    # Возвращаем список в порядке появления & в contract_template.docx
    return [
        get_next_contract_number(),           # 1. № договора (например, 103-ИП)
        get_current_date(),                   # 2. Дата договора
        full_fio,                             # 3. Полное ФИО после "и"
        full_fio,                             # 4. ФИО в реквизитах (Заказчик: &)
        client["Адрес"],                      # 5. Адрес после "РОССИЯ, "
        client["Паспорт (серия и номер)"],    # 6. Паспорт (серия и номер)
        client["Кем выдан"],                  # 7. Кем выдан
        client["Дата выдачи"],                # 8. Дата выдачи паспорта
        client["Код подразделения"],          # 9. Код подразделения
        client["Телефон"],                    # 10. Телефон (в строке e-mail)
        short_fio                             # 11. Подпись: _________ / & М.П.
    ]


def is_contract_exists_for_fio(full_name: str) -> bool:
    """
    Проверяет, существует ли уже договор для указанного ФИО
    :param full_name: Полное ФИО клиента (например, "Иванов Иван Иванович")
    :return: True, если договор уже есть
    """
    try:
        df = pd.read_excel(CONTRACTS_DB_PATH, sheet_name="Registry")
        # Приводим к строке и удаляем лишние пробелы
        df["ФИО"] = df["ФИО"].astype(str).str.strip()

        # Ищем точное совпадение ФИО
        return (df["ФИО"] == full_name.strip()).any()
    except Exception as e:
        logging.warning(f"⚠️ Не удалось проверить дубликат договора: {e}")
        return False  # На всякий случай разрешаем, если ошибка
